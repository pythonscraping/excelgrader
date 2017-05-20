from openpyxl import load_workbook
import glob
from flask import render_template
from flask import request, redirect, url_for
import sqlite3
import os


from flask import Flask
app = Flask(__name__)

UPLOAD_FOLDER = './excelfiles/'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/finalgrading')
def finalgrading():
    finaltext =""
    con = sqlite3.connect('database.db')
    cur = con.cursor()
    cur.execute("select distinct filename FROM formulas")
    filenames = cur.fetchall()
    for file in filenames:
        filename = file[0]
        totalgrade = 0
        totalpoints = 0
        finaltext = finaltext + "File: " + filename + " \n"
        cur.execute("select * FROM formulas WHERE filename=(?)",(filename,))
        couples = cur.fetchall()
        for filename, sheet, cell, value, formula in couples:
            if value is None :
                cur.execute(
                    "select grade FROM uniqueformulas WHERE sheet=:sheet and cell=:cell and formula=:formula and value IS NULL"
                    , {"sheet": sheet, "cell": cell, "formula": formula})
                grade = cur.fetchone()[0]
            else :
                cur.execute("select grade FROM uniqueformulas WHERE sheet=:sheet and cell=:cell and formula=:formula and value=:value"
                            , {"sheet": sheet, "cell": cell, "formula": formula, "value": value})
                grade = cur.fetchone()[0]
            cur.execute(
                "select points FROM cellpoints WHERE sheet=:sheet and cell=:cell"
                , {"sheet": sheet, "cell": cell})
            points = float(cur.fetchone()[0])
            #print (grade*points/100, " / ", points)
            finaltext = finaltext + cell + " from sheet " + str(sheet+1) + " : "
            finaltext = finaltext + str(float(grade)*points/100) + " / " + str(points) + " \n"
            totalgrade = totalgrade + float(grade)*points/100
            totalpoints = totalpoints + points
        finaltext = finaltext + "Total: " + str(totalgrade) + " / " + str(totalpoints) + " \n\n"
    return finaltext.replace("\n","<br>")

@app.route('/hello')
def hello():
    return 'Hello, World'

@app.route('/first')
def first():
    listofexcelfiles = glob.glob("excelfiles/*.xlsx")
    listofsheetnames = []
    for file in listofexcelfiles:
        wb2 = load_workbook(file, read_only = True, data_only=True)
        listofsheetnames = wb2.get_sheet_names()
    return render_template('first.html', number=len(listofsheetnames))

@app.route('/second')
def second():
    listofexcelfiles = glob.glob("excelfiles/*.xlsx")
    con = sqlite3.connect('database.db')
    cur = con.cursor()
    cur.execute("select * from cells ORDER BY sheet ASC")
    rows = cur.fetchall()
    for index, cells in rows:
        print(index, " ", cells)
        for cell in cells.split(','):
            cell = cell.strip()
            cur.execute("INSERT INTO cellpoints (sheet,cell,points) VALUES (?,?,?)",(index, cell, 0))
        for file in listofexcelfiles:
                wb2 = load_workbook(file, read_only = False, data_only=True)
                wb1 = load_workbook(file, read_only = False)
                a = wb2.get_sheet_names()
                #We only deal with the first sheet here
                b =  wb2[ a[index] ]
                b1 =  wb1[ a[index] ]
                file = file.split("/")[1]
                for cell in cells.split(','):
                    cell = cell.strip()
                    print(cell,index," ",file, "TRACKING")
                    try :
                        value = str(b1[cell].value)
                        formula = b[cell].value
                    except :
                        value = ""
                        formula = ""
                    cur.execute("INSERT INTO formulas (filename,sheet,cell,formula,value) VALUES (?,?,?,?,?)",
                        (file,index,cell,formula,value) )
    con.commit()
    con.close()
    return redirect(url_for('third'))

@app.route('/third')
def third():
    con = sqlite3.connect('database.db')
    cur = con.cursor()
    cur.execute("select * from cells ORDER BY sheet ASC")
    rows = cur.fetchall()
    for index, cells in rows:
        for cell in cells.split(','):
            cell = cell.strip()
            cur.execute("SELECT formula,value FROM formulas WHERE sheet=:sheet and cell=:cell", {"sheet": index, "cell": cell})
            couple = cur.fetchall()
            print(couple)
            d = {x: couple.count(x) for x in couple}
            for y, freq in d.items():
                print(y[0],y[1])
                print(freq)
                cur.execute("INSERT INTO uniqueformulas (sheet,cell,formula,value,frequency) VALUES (?,?,?,?,?)",(index, cell, y[1], y[0],freq))
    con.commit()
    cur.close()
    con.close()
    return redirect(url_for('fourth'))

@app.route('/fourth')
def fourth():
    dict={}
    con = sqlite3.connect('database.db')
    cur = con.cursor()
    cur.execute("select rowid,sheet,cell,points from cellpoints ORDER BY sheet ASC, cell ASC")
    pointslist = cur.fetchall()
    for id, sheet, cell, points in pointslist:
        dict[sheet,cell,"points"] = points
        dict[sheet, cell, "points","id"] = id
        cur.execute("SELECT rowid,formula,value,frequency,grade FROM uniqueformulas WHERE sheet=:sheet and cell=:cell ORDER BY frequency DESC",
                    {"sheet": sheet, "cell": cell})
        unique = cur.fetchall()
        dict[sheet, cell, "points","unique"]=unique
        print(unique, "UNIQUE")
    print(dict, "COUCOU")
    cur.execute("select * from uniqueformulas ORDER BY frequency DESC")
    unique = cur.fetchall()
    #print(unique)
    cur.close()
    con.close()
    return render_template('grading.html', dict=dict, pointslist=pointslist)



#Display the list of files
@app.route('/excel')
def excel():
    print('test')
    listofexcelfiles = glob.glob("excelfiles/*.xlsx")
    return render_template('excel.html', files=listofexcelfiles)


#Display the list of cells
@app.route('/displayCells',methods = ['POST', 'GET'])
def displayCells():
    con = sqlite3.connect('database.db')
    cur = con.cursor()
    cur.execute("select * from cells ORDER BY sheet ASC")
    rows = cur.fetchall()
    cur.close()
    con.close()
    return render_template("result2.html",result = rows)

#Handle the POST of listofcells to grade
@app.route('/listofcells',methods = ['POST', 'GET'])
def result():
   if request.method == 'POST':
      con = sqlite3.connect('database.db')
      result1 = request.form
      result = result1.copy()
      #Handling ranges
      for key,value in result.items():
        a = value.split(",")
        for index,element in enumerate(a) :
            if ":" in element:
                b = element.split(":")
                fst = b[0]
                snd = b[1]
                newElement = fst
                for i in range(int(fst[1:])+1,int(snd[1:])+1):
                    print (" THIS IS I :", i)
                    newElement = newElement +","+fst[0]+str(i)
                print(newElement, "NEW ELEMENT")
                a[index]=newElement
        result[key] = ",".join(a)
      for key,value in result.items():
        try:
            cur = con.cursor()
            if len(value)>0:
                cur.execute("INSERT INTO cells (sheet,listofcells) VALUES (?,?)",(key,value) )
                print("Inserting a list")
        except:
            con.rollback()
            print("Problem")
      con.commit()
      cur.close()
      con.close()
      return render_template("result.html",result = result)

@app.route('/updategrades',methods = ['POST', 'GET'])
def updategrades():
    if request.method == 'POST':
        con = sqlite3.connect('database.db')
        cur = con.cursor()
        result = request.form
        for key,value in result.items():
            print(key, " ", value)
            if "grade_" in key:
                print("GRADE")
                rowid = key.split("grade_")[1]
                cur.execute("UPDATE uniqueformulas SET grade=:grade WHERE rowid=:rowid",
                    {"grade": value, "rowid": rowid})
            elif "point_" in key:
                print("POINT")
                rowid = key.split("point_")[1]
                cur.execute("UPDATE cellpoints SET points=:grade WHERE rowid=:rowid",
                            {"grade": value, "rowid": rowid})
            else:
                pass
        con.commit()
        cur.close()
        con.close()
    return redirect(url_for('fourth'))


@app.route('/reset')
def createTables():
    conn = sqlite3.connect('database.db')
    print ("Opened database successfully")
    try:
        conn.execute('DROP TABLE cells')
        conn.execute('DROP TABLE formulas')
        conn.execute('DROP TABLE uniqueformulas')
        conn.execute('DROP TABLE cellpoints')
    except:
        pass
    conn.execute('CREATE TABLE cells (sheet INT, listofcells TEXT)')
    conn.execute('CREATE TABLE formulas (filename TEXT, sheet INT, cell TEXT, formula TEXT, value TEXT)')
    conn.execute('CREATE TABLE uniqueformulas (sheet INT, cell TEXT, formula TEXT, value TEXT, frequency INT,grade INT DEFAULT 0)')
    conn.execute('CREATE TABLE cellpoints (sheet INT, cell TEXT, points REAL DEFAULT NULL)')
    print ("Table created successfully")
    conn.close()
    return redirect(url_for('first'))


@app.route("/upload", methods=["POST"])
def upload():
    uploaded_files = request.files.getlist("file[]")
    print (uploaded_files)
    for file in uploaded_files:
        file.save(os.path.join(UPLOAD_FOLDER, file.filename))
    return ""