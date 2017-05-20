from openpyxl import load_workbook


index = 0
file = "alain.xlsx"
wb2 = load_workbook(file,read_only=False,data_only=True)
wb1 = load_workbook(file,read_only=False)
a = wb2.get_sheet_names()

#We only deal with the first sheet here
b =  wb2[ a[index] ]
b1 =  wb1[ a[index] ]